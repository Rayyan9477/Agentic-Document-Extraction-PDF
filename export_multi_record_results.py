"""
Multi-record Excel and Markdown exporter.
Exports multiple patient records with one record per row.
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_multi_record_to_excel(data: dict, output_path: str):
    """
    Export multi-record extraction to Excel.
    Creates sheets for: All Records, Summary, Record Details
    """
    print(f"\n[Exporting {data['total_records_detected']} records to Excel]")
    
    wb = Workbook()
    
    # Remove default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Sheet 1: All Records Summary (One row per patient)
    ws_summary = wb.create_sheet("All Records")
    
    # Headers
    headers = [
        "Record #", "Patient Name", "Sex", "Patient ID", "DOB", 
        "Referring Physician", "Sedation", "Indications", 
        "Findings Count", "ICD Codes", "CPT Codes", "Plan", "Confidence"
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws_summary.row_dimensions[1].height = 25
    ws_summary.freeze_panes = "A2"
    
    # Data rows
    for row_idx, record in enumerate(data['extracted_records'], start=2):
        fields = record['fields']
        
        # Format findings
        findings = fields.get('findings', [])
        if isinstance(findings, list):
            findings_count = len(findings)
            findings_text = "\n".join(findings[:3])  # Show first 3
            if len(findings) > 3:
                findings_text += f"\n... ({len(findings) - 3} more)"
        else:
            findings_count = 0
            findings_text = str(findings)
        
        # Format ICD codes
        icd_codes = fields.get('icd_codes', [])
        if isinstance(icd_codes, list):
            icd_text = "\n".join(icd_codes)
        else:
            icd_text = str(icd_codes)
        
        # Format CPT codes
        cpt_codes = fields.get('cpt_codes', [])
        if isinstance(cpt_codes, list):
            cpt_text = "\n".join(cpt_codes)
        else:
            cpt_text = str(cpt_codes)
        
        # Format plan
        plan = fields.get('plan', [])
        if isinstance(plan, list):
            plan_text = "\n".join(plan)
        else:
            plan_text = str(plan)
        
        row_data = [
            record['record_id'],
            fields.get('patient_name', ''),
            fields.get('patient_sex', ''),
            fields.get('patient_id', ''),
            fields.get('patient_dob', ''),
            fields.get('referring_physician', ''),
            fields.get('sedation', ''),
            fields.get('indications', ''),
            findings_count,
            icd_text,
            cpt_text,
            plan_text,
            f"{record.get('confidence', 0.0):.0%}"
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        ws_summary.row_dimensions[row_idx].height = 60
    
    # Auto-size columns
    column_widths = {
        1: 10,   # Record #
        2: 25,   # Patient Name
        3: 8,    # Sex
        4: 15,   # Patient ID
        5: 12,   # DOB
        6: 25,   # Referring Physician
        7: 25,   # Sedation
        8: 35,   # Indications
        9: 12,   # Findings Count
        10: 30,  # ICD Codes
        11: 35,  # CPT Codes
        12: 35,  # Plan
        13: 12,  # Confidence
    }
    
    for col_idx, width in column_widths.items():
        ws_summary.column_dimensions[chr(64 + col_idx)].width = width
    
    # Enable auto-filter
    ws_summary.auto_filter.ref = ws_summary.dimensions
    
    # Sheet 2: Detailed Findings (One finding per row)
    ws_findings = wb.create_sheet("Detailed Findings")
    
    headers_findings = ["Record #", "Patient Name", "Finding #", "Finding Text", "Anatomical Location"]
    
    for col_idx, header in enumerate(headers_findings, start=1):
        cell = ws_findings.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws_findings.freeze_panes = "A2"
    
    finding_row = 2
    for record in data['extracted_records']:
        fields = record['fields']
        findings = fields.get('findings', [])
        
        if isinstance(findings, list):
            for finding_idx, finding in enumerate(findings, start=1):
                # Extract location from finding text (in parentheses)
                location = ""
                if "(" in finding and ")" in finding:
                    location = finding.split("(")[-1].split(")")[0]
                
                row_data = [
                    record['record_id'],
                    fields.get('patient_name', ''),
                    finding_idx,
                    finding,
                    location
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_findings.cell(row=finding_row, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = thin_border
                
                finding_row += 1
    
    ws_findings.column_dimensions['A'].width = 10
    ws_findings.column_dimensions['B'].width = 25
    ws_findings.column_dimensions['C'].width = 10
    ws_findings.column_dimensions['D'].width = 60
    ws_findings.column_dimensions['E'].width = 25
    
    # Sheet 3: Processing Summary
    ws_meta = wb.create_sheet("Processing Summary")
    
    meta_data = [
        ("PDF Path", data.get('pdf_path', '')),
        ("Page Number", data.get('page_number', '')),
        ("Detection Method", data.get('detection_method', '')),
        ("Total Records Detected", data.get('total_records_detected', 0)),
        ("Total Records Extracted", len(data.get('extracted_records', []))),
        ("Total VLM Calls", data.get('total_vlm_calls', 0)),
        ("Total Processing Time (s)", f"{data.get('total_processing_time_ms', 0) / 1000:.2f}"),
        ("Average Time per Record (s)", f"{data.get('total_processing_time_ms', 0) / 1000 / max(len(data.get('extracted_records', [])), 1):.2f}"),
    ]
    
    ws_meta.cell(row=1, column=1, value="Property").font = header_font
    ws_meta.cell(row=1, column=1).fill = header_fill
    ws_meta.cell(row=1, column=2, value="Value").font = header_font
    ws_meta.cell(row=1, column=2).fill = header_fill
    
    for row_idx, (prop, value) in enumerate(meta_data, start=2):
        ws_meta.cell(row=row_idx, column=1, value=prop).font = Font(bold=True)
        ws_meta.cell(row=row_idx, column=2, value=value)
    
    ws_meta.column_dimensions['A'].width = 30
    ws_meta.column_dimensions['B'].width = 40
    
    # Save workbook
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"[OK] Excel file created: {output_path}")
    print(f"     File size: {output_path.stat().st_size:,} bytes")
    print(f"     Sheets: {len(wb.sheetnames)}")
    
    return output_path


def export_multi_record_to_markdown(data: dict, output_path: str):
    """Export multi-record extraction to Markdown."""
    print(f"\n[Exporting {data['total_records_detected']} records to Markdown]")
    
    lines = []
    
    # Title
    lines.append("# Multi-Patient Record Extraction Report")
    lines.append("")
    lines.append(f"**Document**: {data.get('pdf_path', 'N/A')}")
    lines.append(f"**Page**: {data.get('page_number', 'N/A')}")
    lines.append(f"**Records Detected**: {data.get('total_records_detected', 0)}")
    lines.append(f"**Processing Time**: {data.get('total_processing_time_ms', 0) / 1000:.2f}s")
    lines.append(f"**VLM Calls**: {data.get('total_vlm_calls', 0)}")
    lines.append("")
    
    # Table of Contents
    lines.append("## Table of Contents")
    lines.append("")
    for record in data['extracted_records']:
        record_id = record['record_id']
        patient_name = record['patient_name']
        lines.append(f"- [Record {record_id}: {patient_name}](#record-{record_id})")
    lines.append("")
    
    # Individual Records
    for record in data['extracted_records']:
        record_id = record['record_id']
        patient_name = record['patient_name']
        fields = record['fields']
        confidence = record.get('confidence', 0.0)
        
        lines.append(f"## Record {record_id}")
        lines.append("")
        lines.append(f"**Patient**: {patient_name}")
        lines.append(f"**Confidence**: {confidence:.0%}")
        lines.append(f"**Extraction Time**: {record.get('extraction_time_ms', 0) / 1000:.2f}s")
        lines.append("")
        
        # Demographics
        lines.append("### Patient Demographics")
        lines.append("")
        lines.append(f"- **Name**: {fields.get('patient_name', 'N/A')}")
        lines.append(f"- **Sex**: {fields.get('patient_sex', 'N/A')}")
        lines.append(f"- **Patient ID**: {fields.get('patient_id', 'N/A')}")
        lines.append(f"- **Date of Birth**: {fields.get('patient_dob', 'N/A')}")
        lines.append(f"- **Referring Physician**: {fields.get('referring_physician', 'N/A')}")
        lines.append("")
        
        # Procedure Details
        lines.append("### Procedure Details")
        lines.append("")
        lines.append(f"**Sedation**: {fields.get('sedation', 'N/A')}")
        lines.append("")
        lines.append(f"**Indications**: {fields.get('indications', 'N/A')}")
        lines.append("")
        
        # Findings
        lines.append("### Findings")
        lines.append("")
        findings = fields.get('findings', [])
        if isinstance(findings, list):
            for finding in findings:
                lines.append(f"- {finding}")
        else:
            lines.append(f"{findings}")
        lines.append("")
        
        # Diagnoses
        lines.append("### ICD-10 Codes")
        lines.append("")
        icd_codes = fields.get('icd_codes', [])
        if isinstance(icd_codes, list):
            for code in icd_codes:
                lines.append(f"- {code}")
        else:
            lines.append(f"{icd_codes}")
        lines.append("")
        
        # Procedures
        lines.append("### CPT Codes")
        lines.append("")
        cpt_codes = fields.get('cpt_codes', [])
        if isinstance(cpt_codes, list):
            if cpt_codes:
                for code in cpt_codes:
                    lines.append(f"- {code}")
            else:
                lines.append("*No CPT codes documented*")
        else:
            lines.append(f"{cpt_codes}")
        lines.append("")
        
        # Plan
        lines.append("### Plan")
        lines.append("")
        plan = fields.get('plan', [])
        if isinstance(plan, list):
            for item in plan:
                lines.append(f"- {item}")
        else:
            lines.append(f"{plan}")
        lines.append("")
        lines.append("---")
        lines.append("")
    
    content = "\n".join(lines)
    
    # Write to file
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding='utf-8')
    
    print(f"[OK] Markdown file created: {output_path}")
    print(f"     File size: {output_path.stat().st_size:,} bytes")
    print(f"     Lines: {len(lines)}")
    
    return content


def main():
    print("=" * 70)
    print("Multi-Record Export")
    print("=" * 70)
    
    # Load multi-record extraction results
    results_file = "multi_record_extraction_results.json"
    print(f"\n[Loading: {results_file}]")
    
    with open(results_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    print(f"[OK] Loaded {data['total_records_detected']} records")
    
    # Export to Excel
    excel_output = "multi_record_results.xlsx"
    export_multi_record_to_excel(data, excel_output)
    
    # Export to Markdown
    markdown_output = "multi_record_results.md"
    export_multi_record_to_markdown(data, markdown_output)
    
    print("\n" + "=" * 70)
    print("Export Complete!")
    print("=" * 70)
    print(f"\nFiles created:")
    print(f"  - {excel_output}")
    print(f"  - {markdown_output}")


if __name__ == "__main__":
    main()
