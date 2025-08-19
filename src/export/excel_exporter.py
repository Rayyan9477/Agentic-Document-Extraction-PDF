"""
Excel Export Module
Provides professional Excel export functionality with formatting and multiple sheets.
"""

import logging
import io
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import os

# Excel libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Professional Excel export with multiple sheets and formatting."""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        # Define professional styling
        self.styles = {
            "header": {
                "font": Font(bold=True, color="FFFFFF", size=12),
                "fill": PatternFill("solid", fgColor="366092"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            },
            "subheader": {
                "font": Font(bold=True, color="000000", size=11),
                "fill": PatternFill("solid", fgColor="D6E3F0"),
                "alignment": Alignment(horizontal="left", vertical="center"),
                "border": Border(bottom=Side(style="thin"))
            },
            "data": {
                "font": Font(color="000000", size=10),
                "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
                "border": Border(
                    left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC")
                )
            },
            "currency": {
                "font": Font(color="000000", size=10),
                "alignment": Alignment(horizontal="right", vertical="center"),
                "number_format": "$#,##0.00"
            },
            "date": {
                "font": Font(color="000000", size=10),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "number_format": "MM/DD/YYYY"
            },
            "confidence": {
                "font": Font(color="000000", size=10),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "number_format": "0%"
            }
        }
    
    def export_structured_data(
        self, 
        processing_results: Dict[str, Any], 
        filename: str = None
    ) -> Tuple[bytes, str]:
        """
        Export structured data to Excel workbook with multiple sheets.
        
        Args:
            processing_results: Dictionary of processing results by filename
            filename: Optional output filename
            
        Returns:
            Tuple of (Excel file bytes, suggested filename)
        """
        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add sheets
            self._add_summary_sheet(wb, processing_results)
            self._add_patient_data_sheet(wb, processing_results)
            self._add_validation_report_sheet(wb, processing_results)
            self._add_raw_data_sheet(wb, processing_results)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Medical_Superbill_Export_{timestamp}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()
            excel_buffer.close()
            
            logger.info(f"Excel export completed: {len(excel_bytes)} bytes")
            return excel_bytes, filename
            
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise
    
    def _add_summary_sheet(self, wb: Workbook, processing_results: Dict[str, Any]):
        """Add summary overview sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Calculate summary statistics
        total_files = len(processing_results)
        successful_files = sum(1 for result in processing_results.values() if result.get("success", False))
        total_patients = sum(len(result.get("structured_data", [])) for result in processing_results.values())
        total_processing_time = sum(result.get("processing_metadata", {}).get("total_duration", 0) 
                                   for result in processing_results.values())
        
        # Title
        ws["A1"] = "Medical Superbill Processing Summary"
        ws.merge_cells("A1:E1")
        self._apply_style(ws["A1"], "header")
        
        # Generation info
        ws["A3"] = "Report Generated:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._apply_style(ws["A3"], "subheader")
        
        # Summary metrics
        metrics = [
            ("Total Files Processed", total_files),
            ("Successfully Processed", successful_files),
            ("Processing Success Rate", f"{(successful_files/max(total_files,1)):.1%}"),
            ("Total Patients Found", total_patients),
            ("Total Processing Time", f"{total_processing_time:.1f} seconds"),
            ("Average Time per File", f"{(total_processing_time/max(total_files,1)):.1f} seconds")
        ]
        
        ws["A5"] = "Processing Metrics"
        self._apply_style(ws["A5"], "subheader")
        
        for i, (metric, value) in enumerate(metrics, start=6):
            ws[f"A{i}"] = metric
            ws[f"B{i}"] = value
            self._apply_style(ws[f"A{i}"], "data")
            self._apply_style(ws[f"B{i}"], "data")
        
        # File breakdown
        ws["A13"] = "File Processing Results"
        self._apply_style(ws["A13"], "subheader")
        
        headers = ["Filename", "Status", "Patients", "Fields Extracted", "Processing Time"]
        for i, header in enumerate(headers, start=1):
            ws.cell(row=14, column=i, value=header)
            self._apply_style(ws.cell(row=14, column=i), "header")
        
        row = 15
        for filename, result in processing_results.items():
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value="✅ Success" if result.get("success") else "❌ Failed")
            ws.cell(row=row, column=3, value=len(result.get("structured_data", [])))
            
            # Count extracted fields
            fields_extracted = 0
            for record in result.get("structured_data", []):
                fields_extracted += len([v for v in record.get("extracted_data", {}).values() if v])
            ws.cell(row=row, column=4, value=fields_extracted)
            
            processing_time = result.get("processing_metadata", {}).get("total_duration", 0)
            ws.cell(row=row, column=5, value=f"{processing_time:.1f}s")
            
            # Apply styling
            for col in range(1, 6):
                self._apply_style(ws.cell(row=row, column=col), "data")
            
            row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _add_patient_data_sheet(self, wb: Workbook, processing_results: Dict[str, Any]):
        """Add detailed patient data sheet."""
        ws = wb.create_sheet("Patient Data")
        
        # Collect all unique fields
        all_fields = set()
        for result in processing_results.values():
            for record in result.get("structured_data", []):
                all_fields.update(record.get("extracted_data", {}).keys())
        
        # Common fields first, then alphabetical
        priority_fields = [
            "patient_name", "date_of_birth", "patient_id", "date_of_service",
            "provider_name", "provider_npi", "cpt_codes", "dx_codes",
            "insurance_company", "policy_number", "copay_amount", "total_charges"
        ]
        
        sorted_fields = priority_fields + [f for f in sorted(all_fields) if f not in priority_fields]
        
        # Headers
        headers = ["Source File", "Patient ID", "Confidence"] + [
            field.replace('_', ' ').title() for field in sorted_fields
        ]
        
        for i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=header)
            self._apply_style(ws.cell(row=1, column=i), "header")
        
        # Patient data
        row = 2
        for filename, result in processing_results.items():
            if not result.get("success"):
                continue
            
            for record in result.get("structured_data", []):
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=record.get("patient_id", "Unknown"))
                ws.cell(row=row, column=3, value=record.get("confidence_score", 0))
                self._apply_style(ws.cell(row=row, column=3), "confidence")
                
                extracted_data = record.get("extracted_data", {})
                
                for i, field in enumerate(sorted_fields, start=4):
                    value = extracted_data.get(field)
                    
                    if isinstance(value, list):
                        # Join list items with semicolons
                        display_value = "; ".join(str(item) for item in value if item)
                    elif value is None or value == "":
                        display_value = ""
                    else:
                        display_value = str(value)
                    
                    cell = ws.cell(row=row, column=i, value=display_value)
                    
                    # Apply field-specific formatting
                    if "date" in field.lower():
                        self._apply_style(cell, "date")
                    elif "amount" in field.lower() or "charge" in field.lower() or "copay" in field.lower():
                        if display_value and display_value != "":
                            try:
                                cell.value = float(display_value.replace('$', '').replace(',', ''))
                                self._apply_style(cell, "currency")
                            except:
                                self._apply_style(cell, "data")
                        else:
                            self._apply_style(cell, "data")
                    else:
                        self._apply_style(cell, "data")
                
                row += 1
        
        # Add confidence color scale
        if row > 2:
            confidence_range = f"C2:C{row-1}"
            rule = ColorScaleRule(
                start_type="num", start_value=0, start_color="FF6B6B",
                mid_type="num", mid_value=0.5, mid_color="FFE66D",
                end_type="num", end_value=1, end_color="51CF66"
            )
            ws.conditional_formatting.add(confidence_range, rule)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
    
    def _add_validation_report_sheet(self, wb: Workbook, processing_results: Dict[str, Any]):
        """Add validation and quality report sheet."""
        ws = wb.create_sheet("Validation Report")
        
        # Title
        ws["A1"] = "Data Validation and Quality Report"
        ws.merge_cells("A1:F1")
        self._apply_style(ws["A1"], "header")
        
        # Validation summary
        ws["A3"] = "Validation Summary"
        self._apply_style(ws["A3"], "subheader")
        
        # Collect validation statistics
        total_records = 0
        records_with_errors = 0
        total_errors = 0
        field_error_counts = {}
        
        for result in processing_results.values():
            if not result.get("success"):
                continue
            
            for record in result.get("structured_data", []):
                total_records += 1
                validation_errors = record.get("validation_errors", [])
                if validation_errors:
                    records_with_errors += 1
                    total_errors += len(validation_errors)
                
                # Count missing fields
                missing_fields = record.get("missing_fields", [])
                for field in missing_fields:
                    field_error_counts[field] = field_error_counts.get(field, 0) + 1
        
        # Validation metrics
        metrics = [
            ("Total Records Processed", total_records),
            ("Records with Validation Issues", records_with_errors),
            ("Data Quality Rate", f"{((total_records - records_with_errors) / max(total_records, 1)):.1%}"),
            ("Total Validation Errors", total_errors),
            ("Average Errors per Record", f"{(total_errors / max(total_records, 1)):.1f}")
        ]
        
        for i, (metric, value) in enumerate(metrics, start=4):
            ws[f"A{i}"] = metric
            ws[f"B{i}"] = value
            self._apply_style(ws[f"A{i}"], "data")
            self._apply_style(ws[f"B{i}"], "data")
        
        # Most common missing fields
        if field_error_counts:
            ws["A10"] = "Most Commonly Missing Fields"
            self._apply_style(ws["A10"], "subheader")
            
            headers = ["Field Name", "Missing Count", "Missing Rate"]
            for i, header in enumerate(headers, start=1):
                ws.cell(row=11, column=i, value=header)
                self._apply_style(ws.cell(row=11, column=i), "header")
            
            sorted_fields = sorted(field_error_counts.items(), key=lambda x: x[1], reverse=True)[:10]
            for i, (field, count) in enumerate(sorted_fields, start=12):
                ws.cell(row=i, column=1, value=field.replace('_', ' ').title())
                ws.cell(row=i, column=2, value=count)
                ws.cell(row=i, column=3, value=f"{(count / max(total_records, 1)):.1%}")
                
                for col in range(1, 4):
                    self._apply_style(ws.cell(row=i, column=col), "data")
        
        # Detailed validation errors
        ws["A25"] = "Detailed Validation Errors"
        self._apply_style(ws["A25"], "subheader")
        
        error_headers = ["Source File", "Patient ID", "Error Type", "Field", "Description"]
        for i, header in enumerate(error_headers, start=1):
            ws.cell(row=26, column=i, value=header)
            self._apply_style(ws.cell(row=26, column=i), "header")
        
        error_row = 27
        for filename, result in processing_results.items():
            if not result.get("success"):
                continue
            
            for record in result.get("structured_data", []):
                patient_id = record.get("patient_id", "Unknown")
                validation_errors = record.get("validation_errors", [])
                
                for error in validation_errors:
                    ws.cell(row=error_row, column=1, value=filename)
                    ws.cell(row=error_row, column=2, value=patient_id)
                    ws.cell(row=error_row, column=3, value="Validation Error")
                    ws.cell(row=error_row, column=4, value="Multiple")
                    ws.cell(row=error_row, column=5, value=str(error))
                    
                    for col in range(1, 6):
                        self._apply_style(ws.cell(row=error_row, column=col), "data")
                    
                    error_row += 1
                
                # Add missing field errors
                missing_fields = record.get("missing_fields", [])
                for field in missing_fields:
                    ws.cell(row=error_row, column=1, value=filename)
                    ws.cell(row=error_row, column=2, value=patient_id)
                    ws.cell(row=error_row, column=3, value="Missing Field")
                    ws.cell(row=error_row, column=4, value=field.replace('_', ' ').title())
                    ws.cell(row=error_row, column=5, value=f"Required field '{field}' not found in document")
                    
                    for col in range(1, 6):
                        self._apply_style(ws.cell(row=error_row, column=col), "data")
                    
                    error_row += 1
        
        self._auto_adjust_columns(ws)
    
    def _add_raw_data_sheet(self, wb: Workbook, processing_results: Dict[str, Any]):
        """Add raw processing data for debugging."""
        ws = wb.create_sheet("Raw Data")
        
        # Title
        ws["A1"] = "Raw Processing Data (for debugging)"
        ws.merge_cells("A1:D1")
        self._apply_style(ws["A1"], "header")
        
        # Headers
        headers = ["Source File", "Patient ID", "Raw Extracted Text", "Processing Metadata"]
        for i, header in enumerate(headers, start=1):
            ws.cell(row=3, column=i, value=header)
            self._apply_style(ws.cell(row=3, column=i), "header")
        
        row = 4
        for filename, result in processing_results.items():
            if not result.get("success"):
                # Show failed files too
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value="PROCESSING FAILED")
                ws.cell(row=row, column=3, value=result.get("error", "Unknown error"))
                ws.cell(row=row, column=4, value="")
                
                for col in range(1, 5):
                    self._apply_style(ws.cell(row=row, column=col), "data")
                row += 1
                continue
            
            # Patient records
            for record in result.get("patient_records", []):
                ws.cell(row=row, column=1, value=filename)
                ws.cell(row=row, column=2, value=record.get("patient_identifier", "Unknown"))
                
                # Truncate long text for Excel
                raw_text = record.get("extracted_text", "")
                if len(raw_text) > 1000:
                    raw_text = raw_text[:1000] + "... [truncated]"
                ws.cell(row=row, column=3, value=raw_text)
                
                metadata = record.get("metadata", {})
                ws.cell(row=row, column=4, value=str(metadata))
                
                for col in range(1, 5):
                    self._apply_style(ws.cell(row=row, column=col), "data")
                
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def _apply_style(self, cell, style_name: str):
        """Apply predefined style to a cell."""
        if style_name in self.styles:
            style = self.styles[style_name]
            if "font" in style:
                cell.font = style["font"]
            if "fill" in style:
                cell.fill = style["fill"]
            if "alignment" in style:
                cell.alignment = style["alignment"]
            if "border" in style:
                cell.border = style["border"]
            if "number_format" in style:
                cell.number_format = style["number_format"]
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set reasonable bounds
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width


# Global exporter instance
excel_exporter = ExcelExporter() if OPENPYXL_AVAILABLE else None