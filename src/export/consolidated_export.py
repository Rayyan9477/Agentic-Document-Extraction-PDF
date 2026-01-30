"""
Consolidated export for multi-record extraction results.

Provides:
- Cross-page duplicate detection
- Excel export with All Records, Duplicates, Page Summary, Processing Summary sheets
- Markdown export with per-record sections
- JSON export
- Record completeness validation
"""

import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import get_logger
from src.extraction.multi_record import DocumentExtractionResult, ExtractedRecord

logger = get_logger(__name__)


def detect_duplicates(
    records: list[ExtractedRecord],
) -> dict[str, list[int]]:
    """
    Detect duplicate records across pages using primary identifier similarity.

    Returns:
        Dict mapping normalized identifier to list of record indices.
    """
    identifier_groups: dict[str, list[int]] = defaultdict(list)

    for idx, record in enumerate(records):
        normalized = " ".join(record.primary_identifier.lower().strip().split())
        identifier_groups[normalized].append(idx)

    duplicates = {
        ident: indices
        for ident, indices in identifier_groups.items()
        if len(indices) > 1
    }

    if duplicates:
        logger.info(
            "duplicates_detected",
            count=len(duplicates),
            identifiers=list(duplicates.keys()),
        )
    else:
        logger.info("no_duplicates_found")

    return duplicates


def validate_record_completeness(
    record: ExtractedRecord, schema: dict[str, Any]
) -> dict[str, Any]:
    """Validate record completeness against schema."""
    fields = record.fields
    schema_fields = schema.get("fields", [])
    expected_count = len(schema_fields)

    required_fields = [f["field_name"] for f in schema_fields if f.get("required")]
    missing_required = []
    for fname in required_fields:
        val = fields.get(fname)
        if val is None or val == "" or val == []:
            missing_required.append(fname)

    empty_fields = []
    for fname, val in fields.items():
        if val is None or val == "" or val == []:
            empty_fields.append(fname)

    non_empty = len(fields) - len(empty_fields)
    completeness = non_empty / max(expected_count, 1)

    return {
        "is_complete": len(missing_required) == 0,
        "missing_required_fields": missing_required,
        "empty_fields": empty_fields,
        "field_count": len(fields),
        "expected_field_count": expected_count,
        "completeness_score": completeness,
    }


def export_excel(
    result: DocumentExtractionResult,
    output_path: str | Path,
) -> None:
    """
    Export extraction results to a consolidated Excel workbook.

    Sheets:
    - All Records: One row per record with all fields
    - Duplicates: Cross-page duplicate analysis
    - Page Summary: Per-page statistics
    - Processing Summary: Overall extraction metrics
    """
    output_path = Path(output_path)
    logger.info("exporting_excel", output_path=str(output_path))

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    dup_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")

    records = result.records
    schema = result.schema

    # Detect duplicates
    duplicates = detect_duplicates(records)
    duplicate_indices: set[int] = set()
    for indices in duplicates.values():
        duplicate_indices.update(indices)

    # Validate records
    validations = [validate_record_completeness(r, schema) for r in records]

    # ── Sheet 1: All Records ──
    ws = wb.create_sheet("All Records")
    headers = ["Record #", "Page", "Primary ID"]
    for f in schema.get("fields", []):
        headers.append(f.get("display_name", f["field_name"]))
    headers.extend(["Confidence", "Is Duplicate", "Completeness"])

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    for ri, (rec, val) in enumerate(zip(records, validations), 2):
        is_dup = (ri - 2) in duplicate_indices
        row = [rec.record_id, rec.page_number, rec.primary_identifier]

        for f in schema.get("fields", []):
            v = rec.fields.get(f["field_name"], "")
            if isinstance(v, list):
                v = "\n".join(str(x) for x in v)
            row.append(str(v) if v else "")

        row.extend([
            f"{rec.confidence:.0%}",
            "YES" if is_dup else "NO",
            f"{val['completeness_score']:.0%}",
        ])

        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border
            if is_dup:
                cell.fill = dup_fill
            elif ri % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[ri].height = 30

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Duplicates ──
    if duplicates:
        ws_d = wb.create_sheet("Duplicates")
        d_headers = ["Primary Identifier", "Occurrences", "Pages", "Record IDs", "Action"]
        for ci, h in enumerate(d_headers, 1):
            cell = ws_d.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ri = 2
        for ident, indices in duplicates.items():
            pages = [records[i].page_number for i in indices]
            rec_ids = [records[i].record_id for i in indices]
            row = [
                ident.title(),
                len(indices),
                ", ".join(map(str, pages)),
                ", ".join(map(str, rec_ids)),
                "Review and merge if same entity",
            ]
            for ci, v in enumerate(row, 1):
                cell = ws_d.cell(row=ri, column=ci, value=v)
                cell.fill = warn_fill
            ri += 1

        for ci, w in enumerate([30, 12, 15, 15, 40], 1):
            ws_d.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Page Summary ──
    ws_p = wb.create_sheet("Page Summary")
    p_headers = ["Page", "Records", "Avg Confidence", "Unique IDs", "Duplicates"]
    for ci, h in enumerate(p_headers, 1):
        cell = ws_p.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    page_stats: dict[int, dict[str, Any]] = defaultdict(
        lambda: {"records": [], "identifiers": set()}
    )
    for rec in records:
        page_stats[rec.page_number]["records"].append(rec)
        page_stats[rec.page_number]["identifiers"].add(
            rec.primary_identifier.lower()
        )

    ri = 2
    for pn in sorted(page_stats.keys()):
        st = page_stats[pn]
        avg_c = sum(r.confidence for r in st["records"]) / len(st["records"])
        page_rec_indices = [
            i for i, r in enumerate(records) if r.page_number == pn
        ]
        page_dups = sum(1 for i in page_rec_indices if i in duplicate_indices)

        row = [pn, len(st["records"]), f"{avg_c:.0%}", len(st["identifiers"]), page_dups]
        for ci, v in enumerate(row, 1):
            ws_p.cell(row=ri, column=ci, value=v)
        ri += 1

    for ci in range(1, len(p_headers) + 1):
        ws_p.column_dimensions[get_column_letter(ci)].width = 18

    # ── Sheet 4: Processing Summary ──
    ws_s = wb.create_sheet("Processing Summary")
    avg_conf = sum(r.confidence for r in records) / max(len(records), 1)
    summary_rows = [
        ("Document Type", result.document_type),
        ("Entity Type", result.entity_type),
        ("PDF Path", result.pdf_path),
        ("Total Pages", result.total_pages),
        ("Total Records", result.total_records),
        ("Unique Records", result.total_records - len(duplicate_indices)),
        ("Duplicate Records", len(duplicate_indices)),
        (
            "Unique Identifiers",
            len({r.primary_identifier.lower() for r in records}),
        ),
        ("Avg Records/Page", f"{result.total_records / max(result.total_pages, 1):.1f}"),
        ("Avg Confidence", f"{avg_conf:.0%}"),
        ("Processing Time (s)", f"{result.total_processing_time_ms / 1000:.1f}"),
        (
            "Avg Time/Record (s)",
            f"{result.total_processing_time_ms / 1000 / max(result.total_records, 1):.1f}",
        ),
        ("Total VLM Calls", result.total_vlm_calls),
        ("Schema Fields", len(schema.get("fields", []))),
    ]

    ws_s.cell(row=1, column=1, value="Metric").font = header_font
    ws_s.cell(row=1, column=2, value="Value").font = header_font
    ws_s.cell(row=1, column=1).fill = header_fill
    ws_s.cell(row=1, column=2).fill = header_fill

    for ri, (metric, value) in enumerate(summary_rows, 2):
        ws_s.cell(row=ri, column=1, value=metric).font = Font(bold=True)
        ws_s.cell(row=ri, column=2, value=value)

    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 40

    # Save
    wb.save(str(output_path))
    logger.info("excel_exported", path=str(output_path), sheets=len(wb.sheetnames))


def export_markdown(
    result: DocumentExtractionResult,
    output_path: str | Path,
) -> None:
    """
    Export extraction results to a consolidated Markdown report.

    Sections:
    - Document summary
    - Duplicate warnings
    - Page-level summary table
    - Per-record detail sections
    """
    output_path = Path(output_path)
    logger.info("exporting_markdown", output_path=str(output_path))

    records = result.records
    schema = result.schema
    duplicates = detect_duplicates(records)

    lines: list[str] = []

    # Title
    title = result.document_type.replace("_", " ").title()
    lines.append(f"# {title} Extraction Report")
    lines.append("")
    lines.append(f"**Document**: {result.pdf_path}")
    lines.append(f"**Pages**: {result.total_pages}")
    lines.append(f"**Total Records**: {result.total_records}")
    unique_count = result.total_records - sum(
        len(v) for v in duplicates.values()
    ) + len(duplicates)
    lines.append(f"**Unique Records**: {unique_count}")
    lines.append(f"**Entity Type**: {result.entity_type}")
    lines.append(
        f"**Processing Time**: {result.total_processing_time_ms / 1000:.1f}s"
    )
    lines.append(f"**VLM Calls**: {result.total_vlm_calls}")
    lines.append("")

    # Duplicates
    if duplicates:
        lines.append("## Duplicate Records Detected")
        lines.append("")
        lines.append(f"Found {len(duplicates)} entities appearing multiple times:")
        lines.append("")
        for ident, indices in duplicates.items():
            pages = [records[i].page_number for i in indices]
            lines.append(
                f"- **{ident.title()}**: Pages {', '.join(map(str, pages))}"
            )
        lines.append("")

    # Page summary table
    lines.append("## Summary by Page")
    lines.append("")
    lines.append("| Page | Records | Avg Confidence |")
    lines.append("|------|---------|----------------|")

    page_groups: dict[int, list[ExtractedRecord]] = defaultdict(list)
    for rec in records:
        page_groups[rec.page_number].append(rec)

    for pn in sorted(page_groups.keys()):
        pr = page_groups[pn]
        avg_c = sum(r.confidence for r in pr) / len(pr)
        lines.append(f"| {pn} | {len(pr)} | {avg_c:.0%} |")
    lines.append("")

    # All records
    lines.append("## All Records")
    lines.append("")

    for rec in records:
        lines.append(
            f"### Record {rec.record_id} - Page {rec.page_number}"
        )
        lines.append(
            f"**{result.entity_type.title()}**: {rec.primary_identifier}"
        )
        lines.append(f"**Confidence**: {rec.confidence:.0%}")
        lines.append("")

        for f in schema.get("fields", []):
            fname = f["field_name"]
            dname = f.get("display_name", fname)
            val = rec.fields.get(fname, "N/A")

            if isinstance(val, list):
                lines.append(f"**{dname}**:")
                for item in val:
                    lines.append(f"  - {item}")
            else:
                lines.append(f"**{dname}**: {val}")

        lines.append("")
        lines.append("---")
        lines.append("")

    content = "\n".join(lines)
    output_path.write_text(content, encoding="utf-8")
    logger.info("markdown_exported", path=str(output_path))


def export_json(
    result: DocumentExtractionResult,
    output_path: str | Path,
) -> None:
    """Export extraction results to JSON."""
    output_path = Path(output_path)
    data = result.to_dict()
    output_path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    logger.info("json_exported", path=str(output_path))
