"""
Display Excel file structure and contents.
"""

import openpyxl
from pathlib import Path


def show_excel_structure(excel_path: str):
    """Display Excel workbook structure."""
    print("\n" + "=" * 70)
    print("EXCEL FILE STRUCTURE")
    print("=" * 70)
    
    wb = openpyxl.load_workbook(excel_path)
    
    print(f"\nFile: {excel_path}")
    print(f"Size: {Path(excel_path).stat().st_size:,} bytes")
    print(f"Sheets: {len(wb.sheetnames)}")
    print(f"\nSheet Names:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
    
    # Display each sheet's contents
    for sheet_name in wb.sheetnames:
        print("\n" + "-" * 70)
        print(f"SHEET: {sheet_name}")
        print("-" * 70)
        
        ws = wb[sheet_name]
        
        # Display first 20 rows
        max_display_rows = min(20, ws.max_row)
        for row_idx in range(1, max_display_rows + 1):
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    # Truncate long values
                    str_value = str(value)
                    if len(str_value) > 50:
                        str_value = str_value[:47] + "..."
                    row_values.append(str_value)
                else:
                    row_values.append("")
            
            print(f"  Row {row_idx}: {' | '.join(row_values)}")
        
        if ws.max_row > max_display_rows:
            print(f"  ... ({ws.max_row - max_display_rows} more rows)")
    
    wb.close()
    print("\n" + "=" * 70)


def main():
    excel_file = "superbill_page1_results.xlsx"
    show_excel_structure(excel_file)
    print("\nExcel file inspection complete!")


if __name__ == "__main__":
    main()
