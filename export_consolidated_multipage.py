"""
Universal Multi-Page Export with Duplicate Detection.

Handles:
- Cross-page duplicate detection
- Consolidated Excel export (all pages, all records)
- Record validation and quality checks
- Summary statistics
"""

import json
from pathlib import Path
from typing import List, Dict, Any, Set, Tuple
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from difflib import SequenceMatcher


def load_extraction_results(json_path: str) -> Dict[str, Any]:
    """Load universal extraction results."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def detect_duplicates(records: List[Dict[str, Any]]) -> Dict[str, List[int]]:
    """
    Detect duplicate records across pages using primary identifier similarity.
    Returns dict mapping normalized identifier to list of record indices.
    """
    print("\n[Detecting Duplicate Records]")
    
    # Group by normalized primary identifier
    identifier_groups = defaultdict(list)
    
    for idx, record in enumerate(records):
        primary_id = record['primary_identifier']
        # Normalize: lowercase, remove extra spaces
        normalized_id = ' '.join(primary_id.lower().strip().split())
        identifier_groups[normalized_id].append(idx)
    
    # Find duplicates (groups with more than 1 record)
    duplicates = {
        identifier: indices 
        for identifier, indices in identifier_groups.items() 
        if len(indices) > 1
    }
    
    if duplicates:
        print(f"[FOUND {len(duplicates)} duplicate identifiers]")
        for identifier, indices in duplicates.items():
            pages = [records[idx]['page_number'] for idx in indices]
            print(f"  - '{identifier}' appears on pages: {pages}")
    else:
        print("[OK] No duplicates found")
    
    return duplicates


def validate_record_completeness(record: Dict[str, Any], schema: Dict[str, Any]) -> Dict[str, Any]:
    """
    Validate record completeness and data quality.
    """
    validation = {
        'is_complete': True,
        'missing_required_fields': [],
        'empty_fields': [],
        'field_count': len(record['fields']),
        'expected_field_count': len(schema['fields']),
        'completeness_score': 0.0
    }
    
    # Check required fields
    required_fields = [f['field_name'] for f in schema['fields'] if f.get('required', False)]
    fields = record['fields']
    
    for field_name in required_fields:
        value = fields.get(field_name)
        if value is None or value == '' or value == []:
            validation['missing_required_fields'].append(field_name)
            validation['is_complete'] = False
    
    # Check for empty fields
    for field_name, value in fields.items():
        if value is None or value == '' or value == []:
            validation['empty_fields'].append(field_name)
    
    # Calculate completeness score
    non_empty_count = validation['field_count'] - len(validation['empty_fields'])
    validation['completeness_score'] = non_empty_count / validation['expected_field_count']
    
    return validation


def export_consolidated_excel(data: Dict[str, Any], output_path: str):
    """
    Export all pages and records to consolidated Excel workbook.
    """
    print(f"\n[Creating Consolidated Excel Export]")
    
    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    warning_fill = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")
    duplicate_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    records = data['records']
    schema = data['schema']
    
    # Detect duplicates
    duplicates = detect_duplicates(records)
    duplicate_indices = set()
    for indices in duplicates.values():
        duplicate_indices.update(indices)
    
    # Validate records
    validations = []
    for record in records:
        validation = validate_record_completeness(record, schema)
        validations.append(validation)
    
    # Sheet 1: All Records (One row per record)
    ws_all = wb.create_sheet("All Records")
    
    # Determine columns from schema
    headers = ["Record ID", "Page", "Primary ID"]
    for field in schema['fields']:
        headers.append(field['display_name'])
    headers.extend(["Confidence", "Is Duplicate", "Completeness"])
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_all.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws_all.row_dimensions[1].height = 25
    ws_all.freeze_panes = "A2"
    
    # Write data rows
    for row_idx, (record, validation) in enumerate(zip(records, validations), start=2):
        fields = record['fields']
        is_duplicate = (row_idx - 2) in duplicate_indices
        
        row_data = [
            record['record_id'],
            record['page_number'],
            record['primary_identifier']
        ]
        
        # Add field values in schema order
        for field in schema['fields']:
            field_name = field['field_name']
            value = fields.get(field_name, '')
            
            # Convert lists to strings
            if isinstance(value, list):
                value = '\n'.join(str(v) for v in value)
            
            row_data.append(str(value) if value else '')
        
        row_data.extend([
            f"{record['confidence']:.0%}",
            "YES" if is_duplicate else "NO",
            f"{validation['completeness_score']:.0%}"
        ])
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border
            
            # Highlight duplicates
            if is_duplicate:
                cell.fill = duplicate_fill
            # Alternating rows
            elif row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        ws_all.row_dimensions[row_idx].height = 30
    
    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        ws_all.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
    
    ws_all.auto_filter.ref = ws_all.dimensions
    
    # Sheet 2: Duplicates Analysis
    if duplicates:
        ws_dupes = wb.create_sheet("Duplicates")
        
        dupe_headers = ["Primary Identifier", "Occurrences", "Pages", "Record IDs", "Action Needed"]
        for col_idx, header in enumerate(dupe_headers, start=1):
            cell = ws_dupes.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row_idx = 2
        for identifier, indices in duplicates.items():
            pages = [records[idx]['page_number'] for idx in indices]
            record_ids = [records[idx]['record_id'] for idx in indices]
            
            row_data = [
                identifier.title(),
                len(indices),
                ', '.join(map(str, pages)),
                ', '.join(map(str, record_ids)),
                "Review and merge if same entity"
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_dupes.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = warning_fill
            
            row_idx += 1
        
        ws_dupes.column_dimensions['A'].width = 30
        ws_dupes.column_dimensions['B'].width = 12
        ws_dupes.column_dimensions['C'].width = 15
        ws_dupes.column_dimensions['D'].width = 15
        ws_dupes.column_dimensions['E'].width = 40
    
    # Sheet 3: Page Summary
    ws_pages = wb.create_sheet("Page Summary")
    
    page_headers = ["Page", "Records", "Avg Confidence", "Unique IDs", "Duplicates"]
    for col_idx, header in enumerate(page_headers, start=1):
        cell = ws_pages.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Group by page
    page_stats = defaultdict(lambda: {'records': [], 'identifiers': set()})
    for record in records:
        page_num = record['page_number']
        page_stats[page_num]['records'].append(record)
        page_stats[page_num]['identifiers'].add(record['primary_identifier'].lower())
    
    for page_num in sorted(page_stats.keys()):
        stats = page_stats[page_num]
        avg_conf = sum(r['confidence'] for r in stats['records']) / len(stats['records'])
        
        # Count duplicates on this page
        page_record_indices = [i for i, r in enumerate(records) if r['page_number'] == page_num]
        page_dupes = sum(1 for idx in page_record_indices if idx in duplicate_indices)
        
        row_data = [
            page_num,
            len(stats['records']),
            f"{avg_conf:.0%}",
            len(stats['identifiers']),
            page_dupes
        ]
        
        row_idx = page_num + 1
        for col_idx, value in enumerate(row_data, start=1):
            ws_pages.cell(row=row_idx, column=col_idx, value=value)
    
    for col_idx in range(1, len(page_headers) + 1):
        ws_pages.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
    
    # Sheet 4: Processing Summary
    ws_summary = wb.create_sheet("Processing Summary")
    
    summary_data = [
        ("Document Type", data['document_type']),
        ("Entity Type", data['entity_type']),
        ("PDF Path", data['pdf_path']),
        ("Total Pages", data['total_pages']),
        ("Total Records", data['total_records']),
        ("Unique Records", data['total_records'] - len(duplicate_indices)),
        ("Duplicate Records", len(duplicate_indices)),
        ("Unique Identifiers", len(set(r['primary_identifier'].lower() for r in records))),
        ("Avg Records per Page", f"{data['total_records'] / data['total_pages']:.1f}"),
        ("Avg Confidence", f"{sum(r['confidence'] for r in records) / len(records):.0%}"),
        ("Processing Time (s)", f"{data['total_processing_time_ms'] / 1000:.2f}"),
        ("Avg Time per Record (s)", f"{data['total_processing_time_ms'] / 1000 / data['total_records']:.2f}"),
        ("Total VLM Calls", data['total_vlm_calls']),
        ("Schema Fields", len(schema['fields'])),
    ]
    
    ws_summary.cell(row=1, column=1, value="Metric").font = header_font
    ws_summary.cell(row=1, column=2, value="Value").font = header_font
    
    for row_idx, (metric, value) in enumerate(summary_data, start=2):
        ws_summary.cell(row=row_idx, column=1, value=metric).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40
    
    # Save workbook
    wb.save(output_path)
    
    print(f"[OK] Excel file created: {output_path}")
    print(f"     Total Records: {data['total_records']}")
    print(f"     Unique Records: {data['total_records'] - len(duplicate_indices)}")
    print(f"     Duplicates: {len(duplicate_indices)}")
    print(f"     Sheets: {len(wb.sheetnames)}")


def export_consolidated_markdown(data: Dict[str, Any], output_path: str):
    """Export consolidated markdown report."""
    print(f"\n[Creating Consolidated Markdown Report]")
    
    records = data['records']
    schema = data['schema']
    duplicates = detect_duplicates(records)
    
    lines = []
    
    # Title
    lines.append(f"# {data['document_type'].replace('_', ' ').title()} Extraction Report")
    lines.append("")
    lines.append(f"**Document**: {data['pdf_path']}")
    lines.append(f"**Pages**: {data['total_pages']}")
    lines.append(f"**Total Records**: {data['total_records']}")
    lines.append(f"**Unique Records**: {data['total_records'] - sum(len(v) for v in duplicates.values()) + len(duplicates)}")
    lines.append(f"**Entity Type**: {data['entity_type']}")
    lines.append(f"**Processing Time**: {data['total_processing_time_ms'] / 1000:.2f}s")
    lines.append("")
    
    # Duplicates warning
    if duplicates:
        lines.append("## ⚠️ Duplicate Records Detected")
        lines.append("")
        lines.append(f"Found {len(duplicates)} entities appearing multiple times:")
        lines.append("")
        for identifier, indices in duplicates.items():
            pages = [records[idx]['page_number'] for idx in indices]
            lines.append(f"- **{identifier.title()}**: Pages {', '.join(map(str, pages))}")
        lines.append("")
    
    # Summary by page
    lines.append("## Summary by Page")
    lines.append("")
    lines.append("| Page | Records | Avg Confidence |")
    lines.append("|------|---------|----------------|")
    
    page_groups = defaultdict(list)
    for record in records:
        page_groups[record['page_number']].append(record)
    
    for page_num in sorted(page_groups.keys()):
        page_records = page_groups[page_num]
        avg_conf = sum(r['confidence'] for r in page_records) / len(page_records)
        lines.append(f"| {page_num} | {len(page_records)} | {avg_conf:.0%} |")
    
    lines.append("")
    
    # All records
    lines.append("## All Records")
    lines.append("")
    
    for record in records:
        lines.append(f"### Record {record['record_id']} - Page {record['page_number']}")
        lines.append(f"**{data['entity_type'].title()}**: {record['primary_identifier']}")
        lines.append(f"**Confidence**: {record['confidence']:.0%}")
        lines.append("")
        
        fields = record['fields']
        for field in schema['fields']:
            field_name = field['field_name']
            display_name = field['display_name']
            value = fields.get(field_name, 'N/A')
            
            if isinstance(value, list):
                lines.append(f"**{display_name}**:")
                for item in value:
                    lines.append(f"  - {item}")
            else:
                lines.append(f"**{display_name}**: {value}")
        
        lines.append("")
        lines.append("---")
        lines.append("")
    
    content = "\n".join(lines)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)
    
    print(f"[OK] Markdown file created: {output_path}")


def main():
    print("=" * 70)
    print("Consolidated Multi-Page Export with Duplicate Detection")
    print("=" * 70)
    
    # Load results
    results_file = "universal_extraction_test_3pages.json"
    print(f"\n[Loading: {results_file}]")
    
    data = load_extraction_results(results_file)
    
    print(f"[OK] Loaded {data['total_records']} records from {data['total_pages']} pages")
    
    # Export to Excel
    excel_output = "consolidated_all_pages.xlsx"
    export_consolidated_excel(data, excel_output)
    
    # Export to Markdown
    markdown_output = "consolidated_all_pages.md"
    export_consolidated_markdown(data, markdown_output)
    
    print("\n" + "=" * 70)
    print("Export Complete!")
    print("=" * 70)


if __name__ == "__main__":
    main()
